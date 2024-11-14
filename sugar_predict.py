'''
    Create the prediction image
'''

import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d

from model import ESN, RLS, LMS
import openpyxl

# from model import ESN, Tikhonov

from lowpass import butter_lowpass_filter
from kalman import kalman_filter
np.random.seed(seed=0)


# Read data from file
def read_sugar_data(file_name):
    '''
    :input: file_name
    :output: data
    '''
    time=np.empty(0)
    data = np.empty(0)
    with open(file_name, 'r') as f:
        lines = f.readlines()
        for line in lines:
            tmp = line.split()
            time = np.hstack((time, float(tmp[0])))
            data = np.hstack((data, float(tmp[1])))  # 2rd column
    return time, data

def interpolate_data(time, data):
    # Construct interpolation function
    f = interp1d(time, data, kind='linear')

    # Interpolation operation
    interpolated_time = np.linspace(time[0], time[-1], num=(len(time)-1)*5 + 1)
    interpolated_data = f(interpolated_time)

    return interpolated_time, interpolated_data



if __name__ == '__main__':
    # Blood glucose data
    time1, bloodsugar1 = read_sugar_data(file_name='blood_sugar(1).txt')
    time2, bloodsugar2 = read_sugar_data(file_name='blood_sugar(2).txt')
    time3, bloodsugar3 = read_sugar_data(file_name='blood_sugar(5).txt')
    time4, bloodsugar4 = read_sugar_data(file_name='blood_sugar(6).txt')
    time5, bloodsugar5 = read_sugar_data(file_name='blood_sugar(9).txt')
    data1 = bloodsugar1

    # Make 5 min/step to 1 min/step:
    time5, bloodsugar5 = interpolate_data(time5, bloodsugar5)

    # Delete the # at the beginning of the corresponding line if need to use a filter:
    
    # Kalman filter
    process_variance = 1e-3
    measurement_variance = 0.1
    # bloodsugar1 = kalman_filter(bloodsugar1, process_variance, measurement_variance)
    # bloodsugar2 = kalman_filter(bloodsugar2, process_variance, measurement_variance)
    # bloodsugar3 = kalman_filter(bloodsugar3, process_variance, measurement_variance)
    # bloodsugar4 = kalman_filter(bloodsugar4, process_variance, measurement_variance)
    # bloodsugar5 = kalman_filter(bloodsugar5, process_variance, measurement_variance)

    # Low-Pass filter  
    # bloodsugar1 = butter_lowpass_filter(bloodsugar1, cutoff=1 / 50, fs=1 / 5, order=10)
    # bloodsugar2 = butter_lowpass_filter(bloodsugar2, cutoff=1 / 50, fs=1 / 5, order=10)
    # bloodsugar3 = butter_lowpass_filter(bloodsugar3, cutoff=1 / 50, fs=1 / 5, order=10)
    # bloodsugar4 = butter_lowpass_filter(bloodsugar4, cutoff=1 / 50, fs=1 / 5, order=10)
    # bloodsugar5 = butter_lowpass_filter(bloodsugar5, cutoff=1 / 50, fs=1 / 5, order=10)

    # Put all the blood sugar values into an array for operation below:
    bloodsugar = np.array(
        [bloodsugar1[:1984], bloodsugar2[:1984], bloodsugar3[:1984], bloodsugar4[:1984], bloodsugar5[:1984]])



    # Multi-step ahead prediction
    step = 20

    # Delete the corresponding # as needed:

    # Train and test with different data
    # data = bloodsugar[1]
    # if step >= 0:
    #     train_U = (bloodsugar[1][:bloodsugar[1].size - step]).reshape(-1, 1)
    #     train_D = (bloodsugar[1][step:]).reshape(-1, 1)
    #     test_U = (bloodsugar[2][:bloodsugar[2].size - step]).reshape(-1, 1)
    #     test_D = (bloodsugar[2][step:]).reshape(-1, 1)
    # else:
    #     train_U = (bloodsugar[1][-step:]).reshape(-1, 1)
    #     train_D = (bloodsugar[1][:bloodsugar[1].size + step]).reshape(-1, 1)
    #     test_U = (bloodsugar[2][-step:]).reshape(-1, 1)
    #     test_D = (bloodsugar[2][:bloodsugar[2].size + step]).reshape(-1, 1)

    # Train and test with the same data(in this case: bloodsugar5)
    T_train = 8000 # training step
    data = bloodsugar5
    T_test = data.size - T_train
    if step >= 0:
        train_U = data[:T_train].reshape(-1, 1)
        train_D = data[step:T_train+step].reshape(-1, 1)
        test_U = data[T_train:T_train + T_test - step].reshape(-1, 1)
        test_D = data[T_train + step:T_train + T_test].reshape(-1, 1)
    else:
        train_U = data[-step:T_train].reshape(-1, 1)
        train_D = data[:T_train+step].reshape(-1, 1)
        test_U = data[T_train:T_train + T_test].reshape(-1, 1)
        test_D = data[T_train + step:T_train + T_test + step].reshape(-1, 1)



    # ESN model
    N_x = 60  # number of nodes
    model = ESN(train_U.shape[1], train_D.shape[1], N_x, density=0.1, leaking_rate=0.9,
                fb_scale=1,fb_seed=0,input_scale=0.01, rho=0.9,seed=0)


    #online RLS
    train_Y, Wout_size,Wout = model.adapt(train_U, train_D,
                                     RLS(N_x, train_D.shape[0], delta=1e-4,
                                         lam=1, update=1))
    # online LMS
    # train_Y, Wout_size,Wout = model.adapt(train_U, train_D,
    #                                  LMS(N_x, train_D.shape[0], step_size=0.003))

    test_Y = model.RLS_predict(test_U,Wout)
    print('trainY:', train_Y[:,0])
    print('testY:', test_Y[:, 0])

    # Training error
    RMSE = np.sqrt(((train_D - train_Y) ** 2)
                   .mean())
    NRMSE = RMSE / np.sqrt(np.var(train_D))
    print(step, 'step(s) ahead prediction')
    print('Training error: RMSE =', RMSE)
    print('Training error: NRMSE =', NRMSE)

    # Testing error
    RMSE = np.sqrt(((test_D - test_Y) ** 2)
                   .mean())
    NRMSE = RMSE / np.sqrt(np.var(test_D))
    print(step, 'step(s) ahead prediction')
    print('Testing error: RMSE =', RMSE)
    print('Testing error: NRMSE =', NRMSE)

    # Data for graph display
    T_disp = (200, 200) # (last 200 of training data, first 200 of testing data)
    print('disp:',T_disp)
    t_axis = np.arange(0,T_disp[0]+T_disp[1], 1)
    disp_D_step1 = np.concatenate((train_D[train_D.size - T_disp[0]:],
                                   test_D[:T_disp[1]]))
    disp_Y_step1 = np.concatenate((train_Y[train_Y.shape[0] - T_disp[0]:],
                                   test_Y[:T_disp[1]]))
    disp_U_step1 = np.concatenate((train_U[train_U.size - T_disp[0]:],
                                   test_U[:T_disp[1]]))
    
    # Save data to Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i in range(len(train_U[:, 0])):
        sheet.cell(row=i + 1, column=1, value=train_U[:, 0][i])
    for i in range(len(test_U[:, 0])):
        sheet.cell(row=i + 1, column=2, value=test_U[:, 0][i])
    for i in range(len(train_D[:, 0])):
        sheet.cell(row=i + 1, column=4, value=train_D[:, 0][i])
    for i in range(len(test_D[:, 0])):
        sheet.cell(row=i + 1, column=5, value=test_D[:, 0][i])
    for i in range(len(train_Y[:, 0])):
        sheet.cell(row=i + 1, column=7, value=train_Y[:, 0][i])
    for i in range(len(test_Y[:, 0])):
        sheet.cell(row=i + 1, column=8, value=test_Y[:, 0][i])
    workbook.save('data.xlsx')
    
    # Graph display
    plt.rcParams['font.size'] = 12
    fig = plt.figure(figsize=(8, 7))
    plt.subplots_adjust(hspace=0.3)

    # Overall blood glucose
    ax1 = fig.add_subplot(2, 1, 1)
    x=range(0, 250)
    x_scaled = [i / 12 for i in x]
    plt.plot(bloodsugar5, color='k', linewidth=1)
    plt.xlabel('step(/5 min)')
    plt.ylabel('blood glucose(mg/dL)')

    # Prediction graph
    ax2 = fig.add_subplot(2, 1, 2)
    ax2.text(-0.15, 1, '', transform=ax2.transAxes)
    ax2.grid(True)
    plt.plot(t_axis*5, disp_D_step1, color='k', linewidth=1, label='target')
    plt.plot(t_axis*5, disp_Y_step1[:,0], color='red', linestyle='--', linewidth=1, label='prediction')
    plt.plot(t_axis * 5 , disp_U_step1[:, 0], color='b', linestyle=':', linewidth=1, label='filtered data')

    plt.xlabel('time(min)')
    plt.ylabel('glucose prediction(mg/dL)')

    plt.legend(loc='upper center')

    plt.show()

